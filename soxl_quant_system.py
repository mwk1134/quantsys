import requests
import json
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

import sys
import io
from contextlib import redirect_stdout


class SOXLQuantTrader:
    """SOXL 퀀트투자 시스템"""

    
    def load_rsi_reference_data(self, filename: str = "data/weekly_rsi_reference.json") -> dict:
        """
        RSI 참조 데이터 로드 (JSON 형식)
        Args:
            filename: RSI 참조 파일명
        Returns:
            dict: RSI 참조 데이터
        """
        try:
            # PyInstaller 실행파일에서 파일 경로 처리
            if getattr(sys, 'frozen', False):
                # 실행파일로 실행된 경우
                if hasattr(sys, '_MEIPASS'):
                    # PyInstaller의 임시 폴더
                    application_path = sys._MEIPASS
                else:
                    # 일반 실행파일
                    application_path = os.path.dirname(sys.executable)
                file_path = os.path.join(application_path, filename)
            else:
                # 스크립트로 실행된 경우
                file_path = filename
            
            # data 폴더가 없으면 생성
            data_dir = os.path.dirname(file_path)
            if data_dir and not os.path.exists(data_dir):
                os.makedirs(data_dir, exist_ok=True)
                print(f"📁 {data_dir} 폴더 생성 완료")
            
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    rsi_data = json.load(f)
                
                # 메타데이터 출력
                metadata = rsi_data.get('metadata', {})
                total_weeks = metadata.get('total_weeks', 0)
                last_updated = metadata.get('last_updated', 'Unknown')
                
                print(f"📊 RSI 참조 데이터 로드 완료")
                print(f"   - 파일 경로: {file_path}")
                print(f"   - 총 {len(rsi_data)-1}개 연도 데이터 ({total_weeks}주차)")
                print(f"   - 마지막 업데이트: {last_updated}")
                
                return rsi_data
            else:
                print(f"⚠️ RSI 참조 파일이 없습니다: {file_path}")
                return {}
        except Exception as e:
            print(f"❌ RSI 참조 데이터 로드 오류: {e}")
            return {}
    
    def get_rsi_from_reference(self, date: datetime, rsi_data: dict) -> float:
        """
        특정 날짜의 RSI 값을 참조 데이터에서 가져오기 (JSON 형식)
        JSON 파일 전체에서 해당 날짜를 찾는 강력한 검색 로직
        Args:
            date: 확인할 날짜
            rsi_data: RSI 참조 데이터 (JSON)
        Returns:
            float: RSI 값 (없으면 None)
        """
        try:
            if not rsi_data:
                return None
            
            date_str = date.strftime('%Y-%m-%d')
            
            # 1단계: 모든 연도에서 해당 날짜가 포함되는 주차 찾기
            available_years = [y for y in rsi_data.keys() if y != 'metadata']
            available_years.sort(reverse=True)  # 최신 연도부터 검색
            
            for year in available_years:
                if 'weeks' not in rsi_data[year]:
                    continue
                    
                weeks = rsi_data[year]['weeks']
                
                # 해당 날짜가 포함되는 주차 찾기
                for week_data in weeks:
                    start_date = week_data['start']
                    end_date = week_data['end']
                    if start_date <= date_str <= end_date:
                        return float(week_data['rsi'])
            
            # 2단계: 정확한 주차가 없으면 가장 가까운 이전 주차의 RSI 사용
            # 모든 연도의 모든 주차를 날짜순으로 정렬하여 검색
            all_weeks = []
            for year in available_years:
                if 'weeks' not in rsi_data[year]:
                    continue
                for week_data in rsi_data[year]['weeks']:
                    week_data_copy = week_data.copy()
                    week_data_copy['year'] = year
                    all_weeks.append(week_data_copy)
            
            # 종료일 기준으로 정렬
            all_weeks.sort(key=lambda x: x['end'])
            
            # 해당 날짜보다 이전의 가장 가까운 주차 찾기
            for week_data in reversed(all_weeks):
                if week_data['end'] <= date_str:
                    return float(week_data['rsi'])
            
            # 3단계: 그래도 없으면 가장 최근 주차의 RSI 사용
            if all_weeks:
                return float(all_weeks[-1]['rsi'])
            
            return None
        except Exception as e:
            print(f"❌ RSI 참조 데이터 조회 오류: {e}")
            return None
    
    def check_and_update_rsi_data(self, filename: str = "data/weekly_rsi_reference.json") -> bool:
        """
        RSI 참조 데이터가 최신인지 확인하고 필요시 업데이트 (JSON 형식)
        Args:
            filename: RSI 참조 파일명
        Returns:
            bool: 업데이트 성공 여부
        """
        try:
            today = datetime.now()
            
            # PyInstaller 실행파일에서 파일 경로 처리
            if getattr(sys, 'frozen', False):
                # 실행파일로 실행된 경우
                application_path = os.path.dirname(sys.executable)
                file_path = os.path.join(application_path, filename)
            else:
                # 스크립트로 실행된 경우
                file_path = filename
            
            # data 폴더가 없으면 생성
            data_dir = os.path.dirname(file_path)
            if data_dir and not os.path.exists(data_dir):
                os.makedirs(data_dir, exist_ok=True)
                print(f"📁 {data_dir} 폴더 생성 완료")
            
            # 기존 RSI 데이터 로드
            if os.path.exists(file_path):
                #print(f"🔍 JSON 파일 로드 시도: {file_path}")
                with open(file_path, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
                
                # 디버깅: 로드된 데이터 구조 확인
                print(f"✅ JSON 파일 로드 성공!")
                #print(f"   - 파일 크기: {os.path.getsize(file_path)} bytes")
                #print(f"   - 로드된 키들: {list(existing_data.keys())}")
                #print(f"   - 총 연도 수: {len([k for k in existing_data.keys() if k != 'metadata'])}")
                
                # 2024년, 2025년 데이터 확인
                if '2024' in existing_data:
                    print(f"   - 2024년 데이터: {len(existing_data['2024']['weeks'])}주차")
                if '2025' in existing_data:
                    print(f"   - 2025년 데이터: {len(existing_data['2025']['weeks'])}주차")
                
                metadata = existing_data.get('metadata', {})
                last_updated = metadata.get('last_updated', '')
                
                if last_updated:
                    last_update_date = datetime.strptime(last_updated, '%Y-%m-%d')
                    print(f"📅 RSI 참조 데이터 마지막 업데이트: {last_updated}")
                    
                    # 마지막 업데이트가 오늘로부터 1주일 이내면 업데이트 불필요
                    if (today - last_update_date).days <= 7:
                        print("✅ RSI 참조 데이터가 최신 상태입니다.")
                        return True
                    
                    print(f"⚠️ RSI 참조 데이터가 {(today - last_update_date).days}일 전 데이터입니다. 업데이트가 필요합니다.")
                else:
                    print("⚠️ RSI 참조 데이터 메타데이터가 없습니다.")
            else:
                print("⚠️ RSI 참조 파일이 없습니다. 전체 데이터 생성이 필요합니다.")
            
            # 사용자에게 업데이트 확인
            print("\n🔄 RSI 참조 데이터 업데이트가 필요합니다.")
            print("📝 제공해주신 2010년~2025년 RSI 데이터를 모두 추가하시겠습니까?")
            print("   (이 작업은 한 번만 수행하면 됩니다)")
            
            return False
            
        except Exception as e:
            print(f"❌ RSI 데이터 확인 오류: {e}")
            return False
    
    def update_rsi_reference_file(self, filename: str = "data/weekly_rsi_reference.json") -> bool:
        """
        RSI 참조 파일을 최신 데이터로 업데이트 (JSON 형식)
        오늘 날짜까지의 주간 RSI를 자동으로 계산하여 업데이트
        Args:
            filename: RSI 참조 파일명
        Returns:
            bool: 업데이트 성공 여부
        """
        try:
            print("🔄 RSI 참조 데이터 업데이트 중...")
            print("📝 오늘 날짜까지의 주간 RSI를 자동 계산하여 업데이트합니다.")
            
            # PyInstaller 실행파일에서 파일 경로 처리
            if getattr(sys, 'frozen', False):
                # 실행파일로 실행된 경우
                application_path = os.path.dirname(sys.executable)
                file_path = os.path.join(application_path, filename)
            else:
                # 스크립트로 실행된 경우
                file_path = filename
            
            # data 폴더가 없으면 생성
            data_dir = os.path.dirname(file_path)
            if data_dir and not os.path.exists(data_dir):
                os.makedirs(data_dir, exist_ok=True)
                print(f"📁 {data_dir} 폴더 생성 완료")
            
            # 기존 JSON 데이터 로드
            existing_data = {}
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            
            # 현재 연도와 주차 계산
            today = datetime.now()
            current_year = today.strftime('%Y')
            
            # QQQ 데이터 가져오기 (최근 1년)
            print("📊 QQQ 데이터 가져오는 중...")
            qqq_data = self.get_stock_data("QQQ", "1y")
            if qqq_data is None:
                print("❌ QQQ 데이터를 가져올 수 없습니다.")
                return False
            
            # 주간 데이터로 변환
            weekly_data = qqq_data.resample('W-FRI').agg({
                'Open': 'first',
                'High': 'max',
                'Low': 'min',
                'Close': 'last',
                'Volume': 'sum'
            }).dropna()
            
            print(f"📈 주간 데이터 {len(weekly_data)}주 계산 완료")
            
            # 현재 연도 데이터 초기화
            if current_year not in existing_data:
                existing_data[current_year] = {
                    "description": f"{current_year}년 주간 RSI 데이터",
                    "weeks": []
                }
            
            # 최근 12주 RSI 계산 및 업데이트
            recent_weeks = weekly_data.tail(12)  # 최근 12주
            
            for i, (week_end, week_row) in enumerate(recent_weeks.iterrows()):
                # 해당 주의 시작일 계산 (월요일)
                week_start = week_end - timedelta(days=4)  # 금요일에서 4일 전 = 월요일
                
                # 주차 번호 계산 (해당 연도의 몇 번째 주인지)
                week_num = week_start.isocalendar()[1]
                
                # RSI 계산
                data_until_week = qqq_data[qqq_data.index <= week_end]
                if len(data_until_week) >= 20:  # 충분한 데이터가 있을 때
                    rsi_value = self.calculate_weekly_rsi(data_until_week)
                    if rsi_value is not None:
                        # 기존 데이터에서 해당 주차 찾기
                        week_exists = False
                        for j, existing_week in enumerate(existing_data[current_year]['weeks']):
                            if existing_week['week'] == week_num:
                                # 기존 데이터 업데이트
                                existing_data[current_year]['weeks'][j] = {
                                    "start": week_start.strftime('%Y-%m-%d'),
                                    "end": week_end.strftime('%Y-%m-%d'),
                                    "week": week_num,
                                    "rsi": round(rsi_value, 2)
                                }
                                week_exists = True
                                break
                        
                        if not week_exists:
                            # 새로운 주차 데이터 추가
                            existing_data[current_year]['weeks'].append({
                                "start": week_start.strftime('%Y-%m-%d'),
                                "end": week_end.strftime('%Y-%m-%d'),
                                "week": week_num,
                                "rsi": round(rsi_value, 2)
                            })
                        
                        print(f"   주차 {week_num}: {week_start.strftime('%m-%d')} ~ {week_end.strftime('%m-%d')} | RSI: {rsi_value:.2f}")
            
            # 주차별로 정렬
            existing_data[current_year]['weeks'].sort(key=lambda x: x['week'])
            
            # 메타데이터 업데이트
            total_weeks = sum(len(year_data['weeks']) for year, year_data in existing_data.items() if year != 'metadata')
            existing_data['metadata'] = {
                "last_updated": today.strftime('%Y-%m-%d'),
                "total_years": len([k for k in existing_data.keys() if k != 'metadata']),
                "total_weeks": total_weeks,
                "description": "QQQ 주간 RSI 참조 데이터 (14주 Wilder's RSI)"
            }
            
            # JSON 파일로 저장
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, ensure_ascii=False, indent=2)
            
            print("✅ RSI 참조 데이터 업데이트 완료!")
            print(f"   - {current_year}년 데이터 업데이트")
            print(f"   - 총 {total_weeks}개 주차 데이터")
            print(f"   - 마지막 업데이트: {today.strftime('%Y-%m-%d')}")
            
            return True
            
        except Exception as e:
            print(f"❌ RSI 참조 파일 업데이트 오류: {e}")
            return False
    
    def __init__(self, initial_capital: float = 9000):
        """
        초기화
        Args:
            initial_capital: 투자원금 (기본값: 9000달러)
        """
        self.initial_capital = initial_capital

        self.current_mode = None  # RSI 기준에 따라 동적으로 결정
        
        # 미국 주식 시장 휴장일 목록 (2024-2025)
        self.us_holidays = [
            # 2024년 휴장일
            "2024-01-01",  # New Year's Day
            "2024-01-15",  # Martin Luther King Jr. Day
            "2024-02-19",  # Washington's Birthday
            "2024-03-29",  # Good Friday
            "2024-05-27",  # Memorial Day
            "2024-06-19",  # Juneteenth
            "2024-07-04",  # Independence Day
            "2024-09-02",  # Labor Day
            "2024-11-28",  # Thanksgiving Day
            "2024-12-25",  # Christmas Day
            
            # 2025년 휴장일
            "2025-01-01",  # New Year's Day
            "2025-01-20",  # Martin Luther King Jr. Day
            "2025-02-17",  # Washington's Birthday
            "2025-04-18",  # Good Friday
            "2025-05-26",  # Memorial Day
            "2025-06-19",  # Juneteenth
            "2025-07-04",  # Independence Day
            "2025-09-01",  # Labor Day
            "2025-11-27",  # Thanksgiving Day
            "2025-12-25",  # Christmas Day
            
            # 특별 휴장일
            "2025-01-09",  # Jimmy Carter National Day of Mourning
        ]
        
        # RSI 참조 데이터 확인 및 업데이트
        if not self.check_and_update_rsi_data():
            print("📊 RSI 참조 데이터 업데이트 중...")
            if self.update_rsi_reference_file():
                print("✅ RSI 참조 데이터 업데이트 완료")
            else:
                print("❌ RSI 참조 데이터 업데이트 실패")
        
        # SF모드 설정
        self.sf_config = {

            "buy_threshold": 3.5,   # 전일 종가 대비 +3.5%에 매수 (매수가)
            "sell_threshold": 1.4,  # 전일 종가 대비 +1.4%에 매도 (매도가)
            "max_hold_days": 30,    # 최대 보유기간 30일
            
            "split_count": 7,       # 7회 분할매수
            "split_ratios": [0.049, 0.127, 0.230, 0.257, 0.028, 0.169, 0.140]
        }
        
        # AG모드 설정 (나중에 사용)
        self.ag_config = {

            "buy_threshold": 3.6,   # 전일 종가 대비 +3.6%에 매수 (매수가)
            "sell_threshold": 3.5,  # 전일 종가 대비 +3.5%에 매도 (매도가)
            "max_hold_days": 7,     # 최대 보유기간 7일
            "split_count": 8,       # 8회 분할매수
            "split_ratios": [0.062, 0.134, 0.118, 0.148, 0.150, 0.182, 0.186, 0.020]
        }
        
        # 포지션 관리 (회차별)
        self.positions = []  # [{"round": 1, "buy_date": date, "buy_price": price, "shares": shares, "amount": amount}]
        self.current_round = 1
        self.available_cash = initial_capital
        

        # 투자원금 관리 (10거래일마다 업데이트)
        self.current_investment_capital = initial_capital
        self.trading_days_count = 0  # 거래일 카운터
        
        # 세션 상태: 사용자 입력 시작일 (파일 저장 없음)
        self.session_start_date: Optional[str] = None
        
        # 테스트용 오늘 날짜 오버라이드 (YYYY-MM-DD)
        self.test_today_override: Optional[str] = None

    def set_test_today(self, date_str: Optional[str]):
        """테스트용 오늘 날짜 설정/해제. None 또는 빈문자면 해제."""
        if not date_str:
            self.test_today_override = None
            print("🧪 테스트 날짜 해제됨 (실제 오늘 날짜 사용)")
            return
        try:
            # 형식 검증
            _ = datetime.strptime(date_str, "%Y-%m-%d")
            self.test_today_override = date_str
            print(f"🧪 테스트 날짜 설정: {date_str}")
        except ValueError:
            print("❌ 날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식으로 입력해주세요.")

    def get_today_date(self) -> datetime:
        """오버라이드된 오늘 날짜(자정)를 datetime으로 반환."""
        if self.test_today_override:
            d = datetime.strptime(self.test_today_override, "%Y-%m-%d").date()
            return datetime(d.year, d.month, d.day)
        return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    def simulate_from_start_to_today(self, start_date: str, quiet: bool = True) -> Dict:
        """
        시작일부터 최근 거래일까지 시뮬레이션 수행하여 현재 포지션 상태를 맞춘다.
        Args:
            start_date: YYYY-MM-DD 형식 시작일
            quiet: 출력 억제 여부 (기본 True)
        Returns:
            Dict: run_backtest 요약 결과
        """
        latest_trading_day = self.get_latest_trading_day()
        end_date_str = latest_trading_day.strftime('%Y-%m-%d')
        if quiet:
            buf = io.StringIO()
            with redirect_stdout(buf):
                result = self.run_backtest(start_date, end_date_str)
            return result
        else:
            return self.run_backtest(start_date, end_date_str)
    
    def is_market_closed(self, date: datetime) -> bool:
        """
        주식 시장 휴장일 확인
        Args:
            date: 확인할 날짜
        Returns:
            bool: 휴장일이면 True, 거래일이면 False
        """
        # 주말 확인 (토요일=5, 일요일=6)
        if date.weekday() >= 5:
            return True
        
        # 휴장일 확인
        date_str = date.strftime("%Y-%m-%d")
        if date_str in self.us_holidays:
            return True
        
        return False

    def _is_dst_approx(self, dt_utc: datetime) -> bool:
        """미국 서머타임 대략 판별 (3~10월은 DST라고 가정)."""
        return 3 <= dt_utc.month <= 10

    def get_us_eastern_now(self) -> datetime:
        """미국 동부시간(ET) 현재시각 (DST 단순 가정)."""
        if self.test_today_override:
            # 테스트 날짜 기준 정오(12:00) ET로 간주
            d = datetime.strptime(self.test_today_override, "%Y-%m-%d")
            return datetime(d.year, d.month, d.day, 12, 0, 0)
        now_utc = datetime.utcnow()
        offset_hours = 4 if self._is_dst_approx(now_utc) else 5
        return now_utc - timedelta(hours=offset_hours)

    def is_regular_session_closed_now(self) -> bool:
        """정규장(09:30~16:00 ET) 기준으로 오늘 장이 마감됐는지 여부."""
        et_now = self.get_us_eastern_now()
        # 거래일이 아니면(주말/휴장일) '이미 마감'으로 간주
        if not self.is_trading_day(et_now):
            return True
        # 16:00 ET 이후면 마감
        return et_now.hour > 16 or (et_now.hour == 16 and et_now.minute >= 0)
    
    def get_latest_trading_day(self) -> datetime:
        """
        가장 최근 거래일 찾기
        Returns:
            datetime: 가장 최근 거래일
        """
        today = self.get_today_date()
        while self.is_market_closed(today):
            today -= timedelta(days=1)
        return today
        
    def get_stock_data(self, symbol: str, period: str = "1mo") -> Optional[pd.DataFrame]:
        """
        Yahoo Finance API를 통해 주식 데이터 가져오기
        Args:
            symbol: 주식 심볼 (예: "SOXL", "QQQ")
            period: 기간 (1d, 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max)
        Returns:
            DataFrame: 주식 데이터 (Date, Open, High, Low, Close, Volume)
        """
        try:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            # 15y가 지원되지 않으면 10y로 대체
            if period == "15y":
                # 먼저 15y 시도, 실패하면 10y로 대체
                params_list = [{'range': '15y', 'interval': '1d'}, {'range': '10y', 'interval': '1d'}]
            else:
                params_list = [{'range': period, 'interval': '1d'}]
            
            print(f"📊 {symbol} 데이터 가져오는 중...")
            
            # 여러 파라미터 시도
            for i, params in enumerate(params_list):
                try:
                    print(f"   시도 {i+1}/{len(params_list)}: range={params['range']}")
                    response = requests.get(url, headers=headers, params=params, timeout=15)
                    
                    if response.status_code == 200:
                        data = response.json()
                        
                        if 'chart' in data and 'result' in data['chart'] and data['chart']['result']:
                            result = data['chart']['result'][0]
                            
                            if 'timestamp' in result and 'indicators' in result:
                                timestamps = result['timestamp']
                                quote_data = result['indicators']['quote'][0]
                                
                                # DataFrame 생성
                                df_data = {
                                    'Date': [datetime.fromtimestamp(ts) for ts in timestamps],
                                    'Open': quote_data.get('open', [None] * len(timestamps)),
                                    'High': quote_data.get('high', [None] * len(timestamps)),
                                    'Low': quote_data.get('low', [None] * len(timestamps)),
                                    'Close': quote_data.get('close', [None] * len(timestamps)),
                                    'Volume': quote_data.get('volume', [None] * len(timestamps))
                                }
                                
                                df = pd.DataFrame(df_data)
                                df = df.dropna()  # NaN 값 제거
                                df.set_index('Date', inplace=True)
                                
                                print(f"✅ {symbol} 데이터 가져오기 성공! ({len(df)}일치 데이터)")
                                return df
                            else:
                                print(f"   ❌ 차트 데이터 구조 오류")
                        else:
                            print(f"   ❌ 차트 결과 없음")
                    else:
                        print(f"   ❌ HTTP 오류: {response.status_code}")
                        
                except Exception as e:
                    print(f"   ❌ 요청 오류: {e}")
                    
                # 마지막 시도가 아니면 계속
                if i < len(params_list) - 1:
                    print(f"   다음 파라미터로 재시도...")
            
            print(f"❌ {symbol} 모든 파라미터 시도 실패")
            return None
                
        except Exception as e:
            print(f"❌ {symbol} 데이터 가져오기 오류: {e}")
            return None
    
    def get_intraday_last_price(self, symbol: str) -> Optional[Tuple[datetime, float]]:
        """
        분봉(1m) 기준으로 오늘의 최신 가격을 가져온다.
        Returns:
            Optional[Tuple[datetime, float]]: (마지막 시각, 마지막 가격)
        """
        try:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            params = {'range': '1d', 'interval': '1m'}
            response = requests.get(url, headers=headers, params=params, timeout=10)
            if response.status_code != 200:
                return None
            data = response.json()
            result = data.get('chart', {}).get('result', [])
            if not result:
                return None
            result0 = result[0]
            timestamps = result0.get('timestamp') or []
            indicators = result0.get('indicators', {})
            quotes = indicators.get('quote', [])
            if not timestamps or not quotes:
                return None
            closes = quotes[0].get('close') or []
            # 마지막 유효 값 찾기
            for ts, close_val in reversed(list(zip(timestamps, closes))):
                if close_val is not None:
                    ts_dt = datetime.utcfromtimestamp(ts)
                    return ts_dt, float(close_val)
            return None
        except Exception:
            return None


    def calculate_weekly_rsi(self, df: pd.DataFrame, window: int = 14) -> float:
        """

        주간 RSI 계산 (제공된 함수 방식 적용)
        Args:
            df: 일일 주가 데이터

            window: RSI 계산 기간 (기본값: 14)
        Returns:
            float: 최신 주간 RSI 값
        """
        try:
            # 주간 데이터로 변환 (금요일 기준)
            weekly_df = df.resample('W-FRI').agg({
                'Open': 'first',
                'High': 'max',
                'Low': 'min',
                'Close': 'last',
                'Volume': 'sum'
            }).dropna()
            

            # 디버깅: 주간 데이터 확인
            print(f"   주간 데이터 변환 결과:")
            print(f"   - 기간: {weekly_df.index[0].strftime('%Y-%m-%d')} ~ {weekly_df.index[-1].strftime('%Y-%m-%d')}")
            print(f"   - 주간 데이터 수: {len(weekly_df)}주")
            print(f"   - 최근 5주 종가: {weekly_df['Close'].tail(5).values}")
            
            if len(weekly_df) < window + 1:
                print(f"❌ 주간 RSI 계산을 위한 데이터 부족 (필요: {window+1}주, 현재: {len(weekly_df)}주)")
                return None
            

            # 제공된 함수 방식으로 RSI 계산
            delta = weekly_df['Close'].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=window).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
            rs = gain / loss
            rsi = 100 - (100 / (1 + rs))
            

            # 디버깅 정보 출력
            latest_rsi = rsi.iloc[-1]
            print(f"📈 QQQ 주간 RSI: {latest_rsi:.2f}")

            print(f"   데이터 기간: {weekly_df.index[0].strftime('%Y-%m-%d')} ~ {weekly_df.index[-1].strftime('%Y-%m-%d')}")
            print(f"   주간 데이터 수: {len(weekly_df)}주")
            print(f"   최근 3개 RSI: {[f'{x:.2f}' if not np.isnan(x) else 'NaN' for x in rsi.tail(3).values]}")
            
            # 상세 계산 과정 출력
            print(f"   최근 3개 계산 과정:")
            for i in range(-3, 0):
                if i + len(weekly_df) >= 0:
                    date_str = weekly_df.index[i].strftime('%Y-%m-%d')
                    delta_val = delta.iloc[i]
                    gain_val = gain.iloc[i]
                    loss_val = loss.iloc[i]
                    rs_val = rs.iloc[i]
                    rsi_val = rsi.iloc[i]
                    print(f"   {date_str}: delta={delta_val:+.4f}, gain={gain_val:.4f}, loss={loss_val:.4f}, RS={rs_val:.4f}, RSI={rsi_val:.2f}")
            
            return latest_rsi
            
        except Exception as e:
            print(f"❌ 주간 RSI 계산 오류: {e}")
            return None
    

    def determine_mode(self, current_rsi: float, prev_rsi: float, prev_mode: str = "SF") -> str:
        """
        구글스프레드시트 수식 기반 모드 판단
        Args:
            current_rsi: 1주전 RSI (현재 적용할 RSI)
            prev_rsi: 2주전 RSI (이전 RSI)
            prev_mode: 전주 모드
        Returns:
            str: "SF" (안전모드) 또는 "AG" (공세모드)
        """
        # RSI 값이 None인 경우 백테스팅 중단
        if current_rsi is None or prev_rsi is None:
            raise ValueError(f"RSI 데이터가 없습니다. current_rsi: {current_rsi}, prev_rsi: {prev_rsi}")
        
        # 안전모드 조건들 (OR로 연결)
        safe_conditions = [
            # RSI > 65 영역에서 하락 (2주전 RSI > 65이고 2주전 > 1주전)
            prev_rsi > 65 and prev_rsi > current_rsi,
            
            # 40 < RSI < 50에서 하락 (2주전 RSI가 40~50 사이이고 2주전 > 1주전)
            40 < prev_rsi < 50 and prev_rsi > current_rsi,
            
            # RSI가 50 밑으로 하락 (2주전 >= 50이고 1주전 < 50)
            prev_rsi >= 50 and current_rsi < 50
        ]
        
        # 공세모드 조건들 (OR로 연결)
        aggressive_conditions = [
            # RSI가 50 위로 상승 (2주전 < 50이고 2주전 < 1주전이고 1주전 > 50)
            prev_rsi < 50 and prev_rsi < current_rsi and current_rsi > 50,
            
            # 50 < RSI < 60에서 상승 (2주전 RSI가 50~60 사이이고 2주전 < 1주전)
            50 < prev_rsi < 60 and prev_rsi < current_rsi,
            
            # RSI < 35 영역에서 상승 (2주전 < 35이고 2주전 < 1주전)
            prev_rsi < 35 and prev_rsi < current_rsi
        ]
        
        # 안전모드 조건 확인
        if any(safe_conditions):
            return "SF"
        
        # 공세모드 조건 확인
        if any(aggressive_conditions):
            return "AG"
        
        # 조건에 없으면 전주 모드 유지
        return prev_mode
    
    def update_mode(self, qqq_data: pd.DataFrame) -> str:
        """
        QQQ 주간 RSI 기반으로 모드 업데이트
        Args:
            qqq_data: QQQ 주가 데이터
        Returns:
            str: 업데이트된 모드
        """
        try:
            # 주간 RSI 계산
            current_rsi = self.calculate_weekly_rsi(qqq_data)
            if current_rsi is None:
                print("⚠️ RSI 계산 실패, 현재 모드 유지")
                return self.current_mode
            
            # 초기 모드가 없는 경우 RSI 기준으로 결정
            if self.current_mode is None:
                # RSI 50을 기준으로 초기 모드 결정
                if current_rsi >= 50:
                    self.current_mode = "SF"  # 안전모드
                else:
                    self.current_mode = "AG"  # 공세모드
                print(f"🎯 초기 모드 결정: {self.current_mode} (RSI: {current_rsi:.2f})")
                return self.current_mode
            
            # 전주 RSI 계산 (주간 데이터에서)
            weekly_df = qqq_data.resample('W-FRI').agg({
                'Open': 'first',
                'High': 'max',
                'Low': 'min',
                'Close': 'last',
                'Volume': 'sum'
            }).dropna()
            
            if len(weekly_df) < 15:
                print("⚠️ 주간 데이터 부족, 현재 모드 유지")
                return self.current_mode
            
            # 제공된 함수 방식으로 전주 RSI 계산
            delta = weekly_df['Close'].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
            rs = gain / loss
            rsi = 100 - (100 / (1 + rs))
            
            prev_rsi = rsi.iloc[-2] if len(rsi) >= 2 else 50.0
            
            # 모드 결정
            new_mode = self.determine_mode(current_rsi, prev_rsi, self.current_mode)
            
            if new_mode != self.current_mode:
                print(f"🔄 모드 전환: {self.current_mode} → {new_mode}")
                print(f"   현재 RSI: {current_rsi:.2f}, 전주 RSI: {prev_rsi:.2f}")
                self.current_mode = new_mode
            else:
                print(f"📊 현재 모드 유지: {self.current_mode} (RSI: {current_rsi:.2f})")
            
            return self.current_mode
            
        except Exception as e:
            print(f"❌ 모드 업데이트 오류: {e}")
            return self.current_mode
    
    def get_current_config(self) -> Dict:
        """현재 모드에 따른 설정 반환"""
        return self.sf_config if self.current_mode == "SF" else self.ag_config
    
    def calculate_buy_sell_prices(self, current_price: float) -> Tuple[float, float]:
        """
        매수/매도 가격 계산
        Args:
            current_price: 현재 주가 (전일 종가)
        Returns:
            Tuple[float, float]: (매수가격, 매도가격)
        """
        config = self.get_current_config()
        

        # 매수가: 전일 종가 대비 상승한 가격 (매수가 > 종가)
        buy_price = current_price * (1 + config["buy_threshold"] / 100)

        # 매도가: 전일 종가 대비 상승한 가격 (매도가 < 종가)
        sell_price = current_price * (1 + config["sell_threshold"] / 100)
        
        return buy_price, sell_price
    
    def calculate_position_size(self, round_num: int) -> float:
        """
        회차별 매수 금액 계산
        Args:
            round_num: 매수 회차 (1부터 시작)
        Returns:
            float: 해당 회차 매수 금액
        """
        config = self.get_current_config()
        
        if round_num <= len(config["split_ratios"]):
            ratio = config["split_ratios"][round_num - 1]

            # 현재 투자원금 사용 (10거래일마다 업데이트됨)
            amount = self.current_investment_capital * ratio
            return amount
        else:
            return 0.0
    

    def calculate_stop_loss_date(self, buy_date: datetime, max_hold_days: int) -> str:
        """
        거래일 기준 손절예정일 계산 (주말 + 미국증시 휴장일 제외)
        Args:
            buy_date: 매수일
            max_hold_days: 최대 보유 거래일 수
        Returns:
            str: 손절예정일 (MM.DD.(요일) 형식)
        """
        try:
            # 요일을 한글로 변환
            weekdays_korean = ['월', '화', '수', '목', '금', '토', '일']
            
            # 거래일 기준으로 날짜 계산 (주말 + 휴장일 제외)
            current_date = buy_date
            trading_days_count = 0
            
            while trading_days_count < max_hold_days:
                current_date += timedelta(days=1)
                
                # 거래일인지 확인 (주말이 아니고 휴장일이 아닌 경우)
                if self.is_trading_day(current_date):
                    trading_days_count += 1
            
            weekday_korean = weekdays_korean[current_date.weekday()]
            return current_date.strftime(f"%m.%d.({weekday_korean})")
            
        except Exception as e:
            print(f"⚠️ 손절예정일 계산 오류: {e}")
            # 오류 시 기본값 반환
            fallback_date = buy_date + timedelta(days=max_hold_days)
            weekday_korean = weekdays_korean[fallback_date.weekday()]
            return fallback_date.strftime(f"%m.%d.({weekday_korean})")
    
    def is_trading_day(self, date: datetime) -> bool:
        """
        해당 날짜가 거래일인지 확인 (주말 + 미국증시 휴장일 제외)
        Args:
            date: 확인할 날짜
        Returns:
            bool: 거래일이면 True, 아니면 False
        """
        # 주말 확인 (토요일=5, 일요일=6)
        if date.weekday() >= 5:
            return False
        
        # 미국증시 휴장일 확인
        date_str = date.strftime("%Y-%m-%d")
        if date_str in self.us_holidays:
            return False
        
        return True
    
    def can_buy_next_round(self) -> bool:
        """다음 회차 매수 가능 여부 확인"""
        config = self.get_current_config()
        
        # 최대 분할매수 횟수 확인
        if self.current_round > config["split_count"]:
            return False
        
        # 시드 확인
        next_amount = self.calculate_position_size(self.current_round)
        if self.available_cash < next_amount:
            return False
        
        return True
    
    def execute_buy(self, buy_price: float, current_date: datetime) -> bool:
        """
        매수 실행
        Args:
            buy_price: 매수 가격
            current_date: 매수 날짜
        Returns:
            bool: 매수 성공 여부
        """
        if not self.can_buy_next_round():
            return False
        

        # 1회시드 금액 계산
        target_amount = self.calculate_position_size(self.current_round)
        
        # 예수금이 부족한 경우 예수금만큼만 매수
        if target_amount > self.available_cash:
            actual_amount = self.available_cash
        else:
            actual_amount = target_amount
        
        shares = int(actual_amount / buy_price)  # 주식 수 (정수)
        final_amount = shares * buy_price
        
        if final_amount <= 0:
            return False
        
        # 포지션 추가
        position = {
            "round": self.current_round,
            "buy_date": current_date,
            "buy_price": buy_price,
            "shares": shares,

            "amount": final_amount,
            "mode": self.current_mode
        }
        
        self.positions.append(position)

        self.available_cash -= final_amount
        self.current_round += 1  # 매수 성공 시에만 회차 증가
        

        print(f"✅ {self.current_round-1}회차 매수 실행: {shares}주 @ ${buy_price:.2f} (총 ${final_amount:,.0f})")
        
        return True
    
    def check_sell_conditions(self, row: pd.Series, current_date: datetime, prev_close: float) -> List[Dict]:
        """
        매도 조건 확인
        Args:
            row: 당일 주가 데이터 (Open, High, Low, Close)
            current_date: 현재 날짜
            prev_close: 전일 종가
        Returns:
            List[Dict]: 매도할 포지션 리스트
        """
        sell_positions = []
        
        for position in self.positions:
            buy_date = position["buy_date"]

            # 거래일 기준으로 보유기간 계산
            hold_days = 0
            temp_date = buy_date
            while temp_date < current_date:
                temp_date += timedelta(days=1)
                if self.is_trading_day(temp_date):
                    hold_days += 1
            
            # 해당 포지션의 모드 설정 가져오기
            position_config = self.sf_config if position["mode"] == "SF" else self.ag_config
            

            # 해당 포지션의 매수체결가 기준으로 매도가 계산
            position_buy_price = position["buy_price"]
            sell_price = position_buy_price * (1 + position_config["sell_threshold"] / 100)
            
            
            # 1. LOC 매도 조건: 종가가 매도목표가에 도달했을 때 (종가 >= 매도목표가)
            daily_close = row['Close']
            if daily_close >= sell_price:
                sell_positions.append({
                    "position": position,
                    "reason": "목표가 도달",

                    "sell_price": daily_close  # 종가에 매도
                })
            
            # 2. 보유기간 초과 시 매도 (당일 종가에 매도)

            elif hold_days > position_config["max_hold_days"]:
                sell_positions.append({
                    "position": position,
                    "reason": f"보유기간 초과 ({hold_days+1}일)",
                    "sell_price": row['Close']  # 종가에 매도
                })
        
        return sell_positions
    

    def execute_sell(self, sell_info: Dict) -> tuple:
        """
        매도 실행
        Args:
            sell_info: 매도 정보
        Returns:

            tuple: (매도 수익금, 매도된 회차)
        """
        position = sell_info["position"]
        sell_price = sell_info["sell_price"]

        sold_round = position["round"]
        
        proceeds = position["shares"] * sell_price
        profit = proceeds - position["amount"]
        profit_rate = (profit / position["amount"]) * 100
        
        # 포지션 제거
        self.positions.remove(position)
        self.available_cash += proceeds
        

        print(f"✅ {sold_round}회차 매도 실행: {position['shares']}주 @ ${sell_price:.2f}")
        print(f"   매도 사유: {sell_info['reason']}")
        print(f"   수익: ${profit:,.0f} ({profit_rate:+.2f}%)")
        

        return proceeds, sold_round
    
    def get_daily_recommendation(self) -> Dict:
        """
        일일 매매 추천 생성
        Returns:
            Dict: 매매 추천 정보
        """
        print("=" * 60)
        print("🚀 SOXL 퀀트투자 일일 매매 추천")
        print("=" * 60)
        

        # 시장 휴장일 확인
        today = datetime.now()
        is_market_closed = self.is_market_closed(today)
        
        if is_market_closed:
            latest_trading_day = self.get_latest_trading_day()
            if today.weekday() >= 5:
                print(f"📅 주말입니다. 최신 거래일({latest_trading_day.strftime('%Y-%m-%d')}) 데이터를 사용합니다.")
            else:
                print(f"📅 휴장일입니다. 최신 거래일({latest_trading_day.strftime('%Y-%m-%d')}) 데이터를 사용합니다.")
        
        # 1. SOXL 데이터 가져오기
        soxl_data = self.get_stock_data("SOXL", "1mo")
        if soxl_data is None:
            return {"error": "SOXL 데이터를 가져올 수 없습니다."}
        
        # 2. QQQ 데이터 가져오기 (주간 RSI 계산용)
        qqq_data = self.get_stock_data("QQQ", "6mo")  # 충분한 데이터 확보
        if qqq_data is None:
            return {"error": "QQQ 데이터를 가져올 수 없습니다."}
        

        # 3. QQQ 주간 RSI 기반 모드 자동 전환
        self.update_mode(qqq_data)
        
        # QQQ 주간 RSI 계산 (표시용)
        weekly_rsi = self.calculate_weekly_rsi(qqq_data)
        if weekly_rsi is None:
            return {"error": "QQQ 주간 RSI를 계산할 수 없습니다."}
        

        # 4. 최신 SOXL 가격 정보 (최소 2일 데이터 필요)
        if len(soxl_data) < 2:
            return {"error": "데이터가 부족합니다. 최소 2일의 데이터가 필요합니다."}
        
        latest_soxl = soxl_data.iloc[-1]
        current_price = latest_soxl['Close']
        current_date = soxl_data.index[-1]

        
        # 전일 종가 계산
        prev_close = soxl_data.iloc[-2]['Close']
        
        # 5. 매수/매도 가격 계산

        buy_price, sell_price = self.calculate_buy_sell_prices(prev_close)
        
        # 6. 매도 조건 확인

        sell_recommendations = self.check_sell_conditions(latest_soxl, current_date, prev_close)
        
        # 7. 매수 조건 확인
        can_buy = self.can_buy_next_round()
        next_buy_amount = self.calculate_position_size(self.current_round) if can_buy else 0
        
        # 8. 포트폴리오 현황
        total_position_value = sum([pos["shares"] * current_price for pos in self.positions])
        total_invested = sum([pos["amount"] for pos in self.positions])
        unrealized_pnl = total_position_value - total_invested
        
        recommendation = {
            "date": current_date.strftime("%Y-%m-%d"),
            "mode": self.current_mode,
            "qqq_weekly_rsi": weekly_rsi,
            "soxl_current_price": current_price,
            "buy_price": buy_price,
            "sell_price": sell_price,
            "can_buy": can_buy,
            "next_buy_round": self.current_round if can_buy else None,
            "next_buy_amount": next_buy_amount,
            "sell_recommendations": sell_recommendations,
            "portfolio": {
                "positions_count": len(self.positions),
                "total_invested": total_invested,
                "total_position_value": total_position_value,
                "unrealized_pnl": unrealized_pnl,
                "available_cash": self.available_cash,
                "total_portfolio_value": self.available_cash + total_position_value
            }
        }
        
        return recommendation
    
    def print_recommendation(self, rec: Dict):
        """매매 추천 출력"""
        if "error" in rec:
            print(f"❌ 오류: {rec['error']}")
            return
        
        print(f"📅 날짜: {rec['date']}")

        mode_name = "안전모드" if rec['mode'] == "SF" else "공세모드"
        print(f"🎯 모드: {rec['mode']} ({mode_name})")
        print(f"📊 QQQ 주간 RSI: {rec['qqq_weekly_rsi']:.2f}")
        print(f"💰 SOXL 현재가: ${rec['soxl_current_price']:.2f}")
        print()
        
        print("📋 오늘의 매매 추천:")
        print("-" * 40)
        
        # 매수 추천
        if rec['can_buy']:
            print(f"🟢 매수 추천: {rec['next_buy_round']}회차")
            print(f"   매수가: ${rec['buy_price']:.2f} (LOC 주문)")
            print(f"   매수금액: ${rec['next_buy_amount']:,.0f}")
            shares = int(rec['next_buy_amount'] / rec['buy_price'])
            print(f"   매수주식수: {shares}주")
        else:
            if self.current_round > self.get_current_config()["split_count"]:
                print("🔴 매수 불가: 모든 분할매수 완료")
            else:
                print("🔴 매수 불가: 시드 부족")
        
        print()
        
        # 매도 추천
        if rec['sell_recommendations']:
            print(f"🔴 매도 추천: {len(rec['sell_recommendations'])}건")
            for sell_info in rec['sell_recommendations']:
                pos = sell_info['position']
                print(f"   {pos['round']}회차 매도: {pos['shares']}주 @ ${sell_info['sell_price']:.2f}")
                print(f"   매도 사유: {sell_info['reason']}")
        else:
            print("🟡 매도 추천 없음")
        
        print()
        print("💼 포트폴리오 현황:")
        print("-" * 40)
        portfolio = rec['portfolio']
        print(f"보유 포지션: {portfolio['positions_count']}개")
        print(f"투자원금: ${portfolio['total_invested']:,.0f}")
        print(f"평가금액: ${portfolio['total_position_value']:,.0f}")
        print(f"평가손익: ${portfolio['unrealized_pnl']:,.0f} ({(portfolio['unrealized_pnl']/portfolio['total_invested']*100) if portfolio['total_invested'] > 0 else 0:+.2f}%)")
        print(f"현금잔고: ${portfolio['available_cash']:,.0f}")
        print(f"총 자산: ${portfolio['total_portfolio_value']:,.0f}")
        
        print()
        print("📊 보유 포지션 상세:")
        print("-" * 40)
        if self.positions:
            for pos in self.positions:
                hold_days = (datetime.now() - pos['buy_date']).days
                current_value = pos['shares'] * rec['soxl_current_price']
                pnl = current_value - pos['amount']
                pnl_rate = (pnl / pos['amount']) * 100
                
                print(f"{pos['round']}회차: {pos['shares']}주 @ ${pos['buy_price']:.2f} ({hold_days}일 보유)")
                print(f"        평가: ${current_value:,.0f} | 손익: ${pnl:,.0f} ({pnl_rate:+.2f}%)")
        else:
            print("보유 포지션 없음")
    
    def reset_portfolio(self):
        """포트폴리오 초기화 (백테스팅용)"""
        self.positions = []
        self.current_round = 1
        self.available_cash = self.initial_capital

        
        # 투자원금 관리 초기화
        self.current_investment_capital = self.initial_capital
        self.trading_days_count = 0
    
    def check_backtest_starting_state(self, start_date: str, rsi_ref_data: dict) -> dict:
        """
        백테스팅 시작 시점의 상태 확인
        Args:
            start_date: 백테스팅 시작일
            rsi_ref_data: RSI 참조 데이터
        Returns:
            dict: 시작 시점 상태 정보
        """
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            
            # 시작일의 주차와 RSI 확인
            days_until_friday = (4 - start_dt.weekday()) % 7
            if days_until_friday == 0 and start_dt.weekday() != 4:
                days_until_friday = 7
            start_week_friday = start_dt + timedelta(days=days_until_friday)
            
            # 시작 주차의 RSI와 모드 확인
            start_week_rsi = self.get_rsi_from_reference(start_week_friday, rsi_ref_data)
            
            # 1주전, 2주전 RSI 확인
            prev_week_friday = start_week_friday - timedelta(days=7)
            two_weeks_ago_friday = start_week_friday - timedelta(days=14)
            
            prev_week_rsi = self.get_rsi_from_reference(prev_week_friday, rsi_ref_data)
            two_weeks_ago_rsi = self.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)
            
            # 시작 모드 결정
            if prev_week_rsi is not None and two_weeks_ago_rsi is not None:
                start_mode = self.determine_mode(prev_week_rsi, two_weeks_ago_rsi, "SF")
            else:
                print(f"❌ 백테스팅 시작 시점의 RSI 데이터가 없습니다.")
                print(f"   시작 주차 RSI: {start_week_rsi}")
                print(f"   1주전 RSI: {prev_week_rsi}")
                print(f"   2주전 RSI: {two_weeks_ago_rsi}")
                return {
                    "error": f"백테스팅 시작 시점의 RSI 데이터가 없습니다. 1주전: {prev_week_rsi}, 2주전: {two_weeks_ago_rsi}",
                    "start_mode": "SF",
                    "start_round": 1,
                    "start_week_rsi": None,
                    "prev_week_rsi": None,
                    "two_weeks_ago_rsi": None
                }
            
            # 해당 모드에서 몇 회차까지 매수했는지 추정
            # (실제로는 과거 매수 기록이 있어야 정확하지만, 여기서는 간단히 추정)
            estimated_round = 1  # 기본값
            
            print(f"📊 백테스팅 시작 상태:")
            print(f"   - 시작일: {start_date}")
            print(f"   - 시작 주차 RSI: {start_week_rsi:.2f}")
            print(f"   - 1주전 RSI: {prev_week_rsi:.2f}")
            print(f"   - 2주전 RSI: {two_weeks_ago_rsi:.2f}")
            print(f"   - 시작 모드: {start_mode}")
            print(f"   - 시작 회차: {estimated_round}회차")
            
            return {
                "start_mode": start_mode,
                "start_round": estimated_round,
                "start_week_rsi": start_week_rsi,
                "prev_week_rsi": prev_week_rsi,
                "two_weeks_ago_rsi": two_weeks_ago_rsi
            }
            
        except Exception as e:
            print(f"❌ 백테스팅 시작 상태 확인 오류: {e}")
            return {
                "start_mode": "SF",
                "start_round": 1,
                "start_week_rsi": None,
                "prev_week_rsi": None,
                "two_weeks_ago_rsi": None
            }
    
    def run_backtest(self, start_date: str, end_date: str = None) -> Dict:
        """
        백테스팅 실행
        Args:
            start_date: 시작 날짜 (YYYY-MM-DD 형식)
            end_date: 종료 날짜 (None이면 오늘까지)
        Returns:
            Dict: 백테스팅 결과
        """
        print(f"🔄 백테스팅 시작: {start_date} ~ {end_date or '오늘'}")

        
        # RSI 참조 데이터 로드
        rsi_ref_data = self.load_rsi_reference_data()
        
        # 포트폴리오 초기화
        self.reset_portfolio()

        
        # 백테스팅 시작 상태 확인
        starting_state = self.check_backtest_starting_state(start_date, rsi_ref_data)
        
        # RSI 데이터가 없는 경우 백테스팅 중단
        if "error" in starting_state:
            return {"error": starting_state["error"]}
        
        # 시작 모드와 회차 설정
        self.current_mode = starting_state["start_mode"]
        self.current_round = starting_state["start_round"]
        
        print(f"🎯 백테스팅 시작 설정:")
        print(f"   - 모드: {self.current_mode}")
        print(f"   - 회차: {self.current_round}")
        print(f"   - 1회시드 예상: ${self.initial_capital * self.get_current_config()['split_ratios'][self.current_round-1]:,.0f}")
        
        # 날짜 파싱
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d") if end_date else datetime.now()
        except ValueError:
            return {"error": "날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식으로 입력해주세요."}
        
        # 장 마감 전에는 종료일을 확정된 최신 거래일로 강제 보정
        try:
            if not self.is_regular_session_closed_now():
                latest_trading_day = self.get_latest_trading_day().date()
                effective_end_date = min(end_dt.date(), latest_trading_day)
                end_dt = datetime(effective_end_date.year, effective_end_date.month, effective_end_date.day, 23, 59, 59)
        except Exception:
            pass
        

        # 충분한 기간의 데이터 가져오기
        data_start = start_dt - timedelta(days=180)
        

        # SOXL 데이터 가져오기 (2011년부터 데이터 확보)
        period_days = (datetime.now() - data_start).days
        if period_days <= 365:
            period = "1y"
        elif period_days <= 730:
            period = "2y"

        elif period_days <= 1825:  # 5년
            period = "5y"

        elif period_days <= 3650:  # 10년
            period = "10y"
        else:
            period = "15y"  # 15년 (SOXL은 2010년 출시)
            
        soxl_data = self.get_stock_data("SOXL", period)
        if soxl_data is None:
            return {"error": "SOXL 데이터를 가져올 수 없습니다."}
        
        # QQQ 데이터 가져오기
        qqq_data = self.get_stock_data("QQQ", period)
        if qqq_data is None:
            return {"error": "QQQ 데이터를 가져올 수 없습니다."}
        
        # 정규장 미마감이고, 마지막 인덱스 날짜가 오늘이면 무조건 제외 (공급사 조기 생성 일봉 방지)
        try:
            if not self.is_regular_session_closed_now():
                today_date = datetime.now().date()
                if len(soxl_data) > 0 and soxl_data.index.max().date() == today_date:
                    soxl_data = soxl_data[soxl_data.index.date < today_date]
                if len(qqq_data) > 0 and qqq_data.index.max().date() == today_date:
                    qqq_data = qqq_data[qqq_data.index.date < today_date]
        except Exception:
            pass

        # 종료일이 데이터의 마지막 날짜와 같고, 정규장이 아직 마감되지 않았다면 마지막 행 제외
        try:
            if end_date:
                end_d = datetime.strptime(end_date, "%Y-%m-%d").date()
                last_date = soxl_data.index.max().date() if len(soxl_data) > 0 else None
                if last_date and end_d == last_date and not self.is_regular_session_closed_now():
                    soxl_data = soxl_data[soxl_data.index.date < last_date]
                    qqq_data = qqq_data[qqq_data.index.date < last_date]
        except Exception:
            pass
        
        # 백테스팅 기간 데이터 필터링 (기존 방식: 타임스탬프 비교)
        soxl_backtest = soxl_data[soxl_data.index >= start_dt]
        soxl_backtest = soxl_backtest[soxl_backtest.index <= end_dt]
        
        if len(soxl_backtest) == 0:
            return {"error": "해당 기간에 대한 데이터가 없습니다."}
        

        # 매매 기록 저장용 (실제 양식에 맞게)
        daily_records = []  # 일별 기록
        current_week_rsi = starting_state["start_week_rsi"]  # 시작 주차 RSI
        current_mode = starting_state["start_mode"]  # 시작 모드
        current_week = 0  # 현재 주차 (첫 번째 주차 처리 후 1이 됨)
        total_realized_pnl = 0  # 누적 실현손익
        total_invested = 0  # 총 투자금
        cash_balance = self.initial_capital  # 현금 잔고
        
        print(f"📊 총 {len(soxl_backtest)}일 백테스팅 진행...")
        

        # 백테스팅 시작일의 전일 종가 설정
        prev_close = None

        if len(soxl_backtest) > 0:
            # 시작일 전날의 종가를 찾기 위해 전체 데이터에서 검색
            start_date_prev = start_dt - timedelta(days=1)
            prev_data = soxl_data[soxl_data.index <= start_date_prev]
            if len(prev_data) > 0:
                prev_close = prev_data.iloc[-1]['Close']
                print(f"📅 백테스팅 시작 전일 종가: {prev_close:.2f} (날짜: {prev_data.index[-1].strftime('%Y-%m-%d')})")
            else:
                print("⚠️ 백테스팅 시작 전일 데이터를 찾을 수 없습니다.")
        
        current_week_friday = None  # 현재 주차의 금요일
        
        for i, (current_date, row) in enumerate(soxl_backtest.iterrows()):
            current_price = row['Close']
            

            # 거래일 카운터 증가 (거래일인 경우에만)
            if self.is_trading_day(current_date):
                self.trading_days_count += 1
                
                # 10거래일마다 투자원금 업데이트 (10, 20, 30, ... 거래일째)
                if self.trading_days_count % 10 == 0 and self.trading_days_count > 0:
                    # 현재 총자산 계산 (현금 + 보유주식 평가금액)
                    total_shares = sum([pos["shares"] for pos in self.positions])
                    total_assets = self.available_cash + (total_shares * current_price)
                    
                    # 투자원금 업데이트
                    old_capital = self.current_investment_capital
                    self.current_investment_capital = total_assets
                    
                    print(f"💰 투자원금 업데이트: {self.trading_days_count}거래일째 - ${old_capital:,.0f} → ${total_assets:,.0f}")
            
            # 현재 날짜가 속하는 주차의 금요일 계산
            days_until_friday = (4 - current_date.weekday()) % 7  # 금요일(4)까지의 일수
            if days_until_friday == 0 and current_date.weekday() != 4:  # 금요일이 아닌데 계산이 0이면 다음 주 금요일
                days_until_friday = 7
            this_week_friday = current_date + timedelta(days=days_until_friday)
            
            # 새로운 주차인지 확인 (금요일이 바뀌었는지)
            if current_week_friday != this_week_friday:
                current_week_friday = this_week_friday
                
                # 새로운 주차의 RSI 값 가져오기 (해당 주차의 금요일 기준)
                current_week_rsi = self.get_rsi_from_reference(this_week_friday, rsi_ref_data)
                
                # 모드 업데이트 (2주전 RSI와 1주전 RSI 비교)
                # 2주전과 1주전 RSI 계산
                prev_week_friday = this_week_friday - timedelta(days=7)  # 1주전
                two_weeks_ago_friday = this_week_friday - timedelta(days=14)  # 2주전
                
                prev_week_rsi = self.get_rsi_from_reference(prev_week_friday, rsi_ref_data)  # 1주전 RSI
                two_weeks_ago_rsi = self.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)  # 2주전 RSI
                
                # RSI 데이터가 없는 경우 백테스팅 중단
                if prev_week_rsi is None or two_weeks_ago_rsi is None:
                    return {"error": f"RSI 데이터가 없습니다. 1주전 RSI: {prev_week_rsi}, 2주전 RSI: {two_weeks_ago_rsi}"}
                
                # 모드 결정 (2주전 vs 1주전 비교)
                new_mode = self.determine_mode(prev_week_rsi, two_weeks_ago_rsi, current_mode)
                if new_mode != current_mode:
                    prev_rsi_display = f"{prev_week_rsi:.2f}" if prev_week_rsi is not None else "None"
                    two_weeks_rsi_display = f"{two_weeks_ago_rsi:.2f}" if two_weeks_ago_rsi is not None else "None"
                    print(f"🔄 백테스팅 모드 전환: {current_mode} → {new_mode} (1주전 RSI: {prev_rsi_display}, 2주전 RSI: {two_weeks_rsi_display})")
                    print(f"   현재 회차: {self.current_round} → 최대 회차: {7 if new_mode == 'SF' else 8}")
                    current_mode = new_mode
                    self.current_mode = new_mode  # 클래스 변수도 업데이트
                    # 모드 변경 시 current_round 유지 (최대 회차만 변경)
                
                current_week += 1  # 주차 번호 증가 (0 → 1, 1 → 2, ...)
                current_rsi_display = f"{current_week_rsi:.2f}" if current_week_rsi is not None else "None"
                print(f"📅 주차 {current_week}: ~{this_week_friday.strftime('%m-%d')} | RSI: {current_rsi_display}")
            
            # 매매 실행 (전일 종가가 있는 경우만)
            if prev_close is not None:

                # 현재 모드 설정 가져오기
                config = self.sf_config if current_mode == "SF" else self.ag_config
                

                # 매수/매도 가격 계산 (전일 종가 기준)
                buy_price = prev_close * (1 + config["buy_threshold"] / 100)  # 매수가
                sell_price = prev_close * (1 + config["sell_threshold"] / 100)  # 매도가 (임시, 매수 체결 시 재계산됨)
                
                # 매도 조건 확인 및 실행
                sell_recommendations = self.check_sell_conditions(row, current_date, prev_close)

                daily_realized = 0
                sell_date = ""
                sell_executed_price = 0
                
                sold_rounds = []  # 매도된 회차들 추적
                sold_positions = []  # 매도된 포지션들 (매수 행에 기록용)
                
                for sell_info in sell_recommendations:

                    position = sell_info["position"]
                    proceeds, sold_round = self.execute_sell(sell_info)
                    realized_pnl = proceeds - position["amount"]
                    daily_realized += realized_pnl
                    total_realized_pnl += realized_pnl
                    cash_balance += proceeds
                    sold_rounds.append(sold_round)
                    
                    # 매도 정보를 매수 행에 기록하기 위해 저장
                    # 요일을 한글로 변환
                    weekdays_korean = ['월', '화', '수', '목', '금', '토', '일']
                    weekday_korean = weekdays_korean[current_date.weekday()]
                    sold_positions.append({
                        "round": sold_round,
                        "sell_date": current_date.strftime(f"%m.%d.({weekday_korean})"),
                        "sell_price": sell_info["sell_price"],

                        "realized_pnl": realized_pnl
                    })
                
                # 매도된 회차 수만큼 current_round 감소
                if sold_rounds:
                    sold_count = len(sold_rounds)
                    self.current_round = max(1, self.current_round - sold_count)
                    print(f"🔄 매도 완료: {sold_count}개 회차 매도 → current_round: {self.current_round}")
                
                # 매수 조건 확인 및 실행

                buy_executed = False
                buy_price_executed = 0
                buy_quantity = 0
                buy_amount = 0
                current_round_before_buy = self.current_round  # 매수 전 회차 저장
                
                if self.can_buy_next_round():

                    # LOC 매수 조건: 매수가가 종가보다 유리할 때 (매수가 > 종가)
                    daily_close = row['Close']
                    if buy_price > daily_close:
                        if self.execute_buy(daily_close, current_date):  # 종가에 매수
                            buy_executed = True
                            position = self.positions[-1]
                            buy_price_executed = position["buy_price"]
                            buy_quantity = position["shares"]
                            buy_amount = position["amount"]
                            total_invested += buy_amount
                            cash_balance -= buy_amount
                            
                            # 매수 체결 시 매도목표가 재계산 (매수체결된 날의 종가 기준)
                            sell_price = daily_close * (1 + config["sell_threshold"] / 100)
                            
                            # 매수 행에서 매도 정보 초기화 (나중에 매도되면 업데이트됨)
                            sell_date = ""
                            sell_executed_price = 0
                
                # 현재 보유 주식수와 평가손익 계산
                total_shares = sum([pos["shares"] for pos in self.positions])
                position_value = total_shares * current_price
                
                # 보유 주식의 매수 원가 계산
                total_buy_cost = sum([pos["amount"] for pos in self.positions])
                
                
                # 일별 기록 생성
                # 요일을 한글로 변환
                weekdays_korean = ['월', '화', '수', '목', '금', '토', '일']
                weekday_korean = weekdays_korean[current_date.weekday()]
                
                # 매도 정보 초기화 (현재 날짜의 매수 행에는 매도 정보 없음)
                sell_date_final = ""
                sell_executed_price_final = 0
                realized_pnl_final = 0
                
                daily_record = {
                    "date": current_date.strftime(f"%y.%m.%d.({weekday_korean})"),
                    "week": current_week,
                    "rsi": current_week_rsi or 50.0,
                    "mode": current_mode,
                    "current_round": min(current_round_before_buy, 7 if current_mode == "SF" else 8),  # 매수 전 회차 사용 (최대값 제한)
                    "seed_amount": self.calculate_position_size(current_round_before_buy) if buy_executed else 0,
                    "buy_order_price": buy_price,
                    "close_price": current_price,
                    "sell_target_price": sell_price,
                    "stop_loss_date": self.calculate_stop_loss_date(current_date, config["max_hold_days"]),
                    "d": 0,  # D 컬럼 (의미 불명)
                    "trading_days": i + 1,
                    "buy_executed_price": buy_price_executed,
                    "buy_quantity": buy_quantity,
                    "buy_amount": buy_amount,
                    "buy_round": current_round_before_buy if buy_executed else 0,  # 매수 회차 저장
                    "commission": 0.0,
                    "sell_date": sell_date_final,
                    "sell_executed_price": sell_executed_price_final,
                    "holding_days": 0,  # 보유기간 (거래일 기준)
                    "holdings": total_shares,
                    "realized_pnl": realized_pnl_final,
                    "cumulative_realized": total_realized_pnl,
                    "daily_realized": daily_realized,
                    "update": False,
                    "investment_update": self.initial_capital,
                    "withdrawal": False,
                    "withdrawal_amount": 0,
                    "seed_increase": 0,
                    "position_value": position_value,
                    "cash_balance": cash_balance,
                    "total_assets": cash_balance + position_value
                }
                
                daily_records.append(daily_record)
                
                # 오늘 매도된 포지션들의 정보를 과거 매수 행에 기록 (daily_record 생성 후)
                if sold_positions:
                    for sold_pos in sold_positions:
                        
                        # 해당 회차의 매수 행을 찾아서 매도 정보 업데이트
                        found = False
                        for record in daily_records:
                            if (record.get('buy_executed_price', 0) > 0 and 
                                record.get('buy_quantity', 0) > 0 and
                                record.get('sell_date', '') == ''):  # 아직 매도되지 않은 행
                                
                                # 해당 회차인지 확인 (buy_round로 정확한 매칭)
                                if record.get('buy_round', 0) == sold_pos["round"]:
                                    # 보유기간 계산 (거래일 기준)
                                    try:
                                        buy_date_str = record['date']
                                        sell_date_str = sold_pos["sell_date"]
                                        
                                        # 날짜 파싱 (예: "25.01.02.(목)" -> "2025-01-02")
                                        buy_date_str_clean = buy_date_str.split('(')[0].strip().rstrip('.')
                                        sell_date_str_clean = sell_date_str.split('(')[0].strip().rstrip('.')
                                        
                                        buy_date = datetime.strptime(buy_date_str_clean, "%y.%m.%d")
                                        sell_date = datetime.strptime(sell_date_str_clean, "%m.%d")
                                        
                                        # 연도 보정 (매도일에는 연도가 없으므로)
                                        if sell_date.month < buy_date.month or (sell_date.month == buy_date.month and sell_date.day < buy_date.day):
                                            sell_date = sell_date.replace(year=buy_date.year + 1)
                                        else:

                                            sell_date = sell_date.replace(year=buy_date.year)
                                        
                                        # 거래일 계산 (주말 + 휴장일 제외)
                                        holding_days = 0
                                        temp_date = buy_date
                                        while temp_date <= sell_date:
                                            if self.is_trading_day(temp_date):
                                                holding_days += 1
                                            temp_date += timedelta(days=1)
                                        
                                        record['holding_days'] = holding_days
                                        
                                    except Exception as e:
                                        print(f"⚠️ 보유기간 계산 오류: {e}")
                                        record['holding_days'] = 0
                                    
                                    record['sell_date'] = sold_pos["sell_date"]
                                    record['sell_executed_price'] = sold_pos["sell_price"]
                                    record['realized_pnl'] = sold_pos["realized_pnl"]
                                    found = True
                                    break
                        
            
            # 진행상황 출력
            if (i + 1) % 10 == 0:
                print(f"진행: {i+1}/{len(soxl_backtest)}일 ({(i+1)/len(soxl_backtest)*100:.1f}%)")

            
            prev_close = current_price
        
        # 최종 결과 계산

        final_value = daily_records[-1]["total_assets"] if daily_records else self.initial_capital
        total_return = ((final_value - self.initial_capital) / self.initial_capital) * 100
        
        summary = {
            "start_date": start_date,
            "end_date": end_date or datetime.now().strftime("%Y-%m-%d"),

            "trading_days": len(soxl_backtest),
            "initial_capital": self.initial_capital,
            "final_value": final_value,
            "total_return": total_return,
            "final_positions": len(self.positions),

            "daily_records": daily_records
        }

        
        # MDD 계산 및 출력
        mdd_info = self.calculate_mdd(daily_records)
        
        print("✅ 백테스팅 완료!")

        print(f"\n📊 백테스팅 결과 요약:")
        print(f"   📅 기간: {start_date} ~ {end_date or datetime.now().strftime('%Y-%m-%d')}")
        print(f"   💰 초기자본: ${self.initial_capital:,.0f}")
        print(f"   💰 최종자산: ${final_value:,.0f}")
        print(f"   📈 총수익률: {total_return:+.2f}%")
        print(f"   📦 최종보유포지션: {len(self.positions)}개")
        print(f"\n⚠️ 리스크 지표:")
        print(f"   📉 MDD (최대낙폭): {mdd_info.get('mdd_percent', 0.0):.2f}%")
        print(f"   📅 MDD 발생일: {mdd_info.get('mdd_date', '')}")
        print(f"   💰 최저자산: ${mdd_info.get('mdd_value', 0.0):,.0f}")
        print(f"   📅 MDD 발생 최고자산일: {mdd_info.get('mdd_peak_date', '')}")
        print(f"   📅 최고자산일: {mdd_info.get('overall_peak_date', '')}")
        print(f"   💰 최고자산: ${mdd_info.get('overall_peak_value', 0.0):,.0f}")
        
        return summary
    

    
    def get_week_number(self, date: datetime) -> int:
        """날짜로부터 주차 계산"""
        year = date.year
        week_num = date.isocalendar()[1]
        return f"{year}W{week_num:02d}"
    
    def calculate_mdd(self, daily_records: List[Dict]) -> Dict:
        """
        MDD (Maximum Drawdown) 계산
        Args:
            daily_records: 일별 백테스팅 기록
        Returns:
            Dict: MDD 정보
        """
        if not daily_records:
            return {
                "mdd_percent": 0.0, 
                "mdd_date": "", 
                "mdd_value": 0.0, 
                "mdd_peak_date": "",  # MDD 계산 시점의 최고자산일
                "overall_peak_date": "",  # 전체 기간 최고자산일
                "overall_peak_value": 0.0  # 전체 기간 최고자산
            }
        
        max_assets = 0.0
        max_drawdown = 0.0
        mdd_peak_date = ""  # MDD 계산 시점의 최고자산일
        mdd_date = ""
        mdd_value = 0.0
        
        # 전체 기간 최고자산 추적
        overall_max_assets = 0.0
        overall_peak_date = ""
        
        # MDD 계산용 변수들
        current_peak_assets = 0.0
        current_peak_date = ""
        
        for record in daily_records:
            current_assets = record.get('total_assets', 0.0)
            
            # 전체 기간 최고자산 갱신
            if current_assets > overall_max_assets:
                overall_max_assets = current_assets
                overall_peak_date = record.get('date', '')
            
            # 새로운 최고자산 갱신 (MDD 계산용)
            if current_assets > current_peak_assets:
                current_peak_assets = current_assets
                current_peak_date = record.get('date', '')
            
            # 현재 자산이 현재 최고자산보다 낮으면 낙폭 계산
            if current_peak_assets > 0:
                drawdown = (current_peak_assets - current_assets) / current_peak_assets * 100
                if drawdown > max_drawdown:
                    max_drawdown = drawdown
                    mdd_date = record.get('date', '')
                    mdd_value = current_assets
                    mdd_peak_date = current_peak_date  # MDD 발생 시점의 기준 최고자산일
        
        return {
            "mdd_percent": max_drawdown,
            "mdd_date": mdd_date,
            "mdd_value": mdd_value,
            "mdd_peak_date": mdd_peak_date,  # MDD 계산 시점의 최고자산일
            "overall_peak_date": overall_peak_date,  # 전체 기간 최고자산일
            "overall_peak_value": overall_max_assets  # 전체 기간 최고자산
        }
    
    def export_backtest_to_excel(self, backtest_result: Dict, filename: str = None):
        """
        백테스팅 결과를 엑셀 파일로 내보내기
        Args:
            backtest_result: 백테스팅 결과
            filename: 파일명 (None이면 자동 생성)
        """
        if "error" in backtest_result:
            print(f"❌ 엑셀 내보내기 실패: {backtest_result['error']}")
            return
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SOXL_백테스팅_{backtest_result['start_date']}_{timestamp}.xlsx"
        
        # 엑셀 워크북 생성
        wb = openpyxl.Workbook()

        
        # 가운데 정렬 설정
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 요약 시트
        ws_summary = wb.active
        ws_summary.title = "백테스팅 요약"

        
        # 첫 번째 행 고정 (헤더 고정)
        ws_summary.freeze_panes = "A2"
        
        # MDD 계산
        mdd_info = self.calculate_mdd(backtest_result['daily_records'])
        
        # 요약 데이터 작성
        summary_data = [
            ["SOXL 퀀트투자 백테스팅 결과", ""],
            ["", ""],
            ["시작일", backtest_result['start_date']],
            ["종료일", backtest_result['end_date']],
            ["거래일수", f"{backtest_result['trading_days']}일"],
            ["", ""],
            ["초기자본", f"${backtest_result['initial_capital']:,.0f}"],
            ["최종자산", f"${backtest_result['final_value']:,.0f}"],
            ["총수익률", f"{backtest_result['total_return']:+.2f}%"],

            ["최종보유포지션", f"{backtest_result['final_positions']}개"],
            ["", ""],

            ["=== 리스크 지표 ===", ""],
            ["MDD (최대낙폭)", f"{mdd_info.get('mdd_percent', 0.0):.2f}%"],
            ["MDD 발생일", mdd_info.get('mdd_date', '')],
            ["최저자산", f"${mdd_info.get('mdd_value', 0.0):,.0f}"],
            ["MDD 발생 최고자산일", mdd_info.get('mdd_peak_date', '')],
            ["최고자산일", mdd_info.get('overall_peak_date', '')],
            ["최고자산", f"${mdd_info.get('overall_peak_value', 0.0):,.0f}"]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):

            cell1 = ws_summary.cell(row=row_idx, column=1, value=label)
            cell2 = ws_summary.cell(row=row_idx, column=2, value=value)
            cell1.alignment = center_alignment
            cell2.alignment = center_alignment
        
        # 스타일 적용
        title_font = Font(size=16, bold=True)

        title_cell = ws_summary.cell(row=1, column=1)
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        
        # 상세 거래 내역 시트

        ws_detail = wb.create_sheet("매매 상세내역")
        

        # 첫 번째 행 고정 (헤더 고정)
        ws_detail.freeze_panes = "A2"
        
        # 헤더 작성 (실제 양식에 맞게)
        headers = [

            "날짜", "주차", "RSI", "모드", "현재회차", "1회시드", 
            "매수주문가", "종가", "매도목표가", "손절예정일", "거래일수", 
            "매수체결", "수량", "매수대금", "매도일", "매도체결", "보유기간",
            "보유", "실현손익", "누적실현", "당일실현",
            "예수금", "총자산"
        ]
        
        header_font = Font(size=11, bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

            cell.alignment = center_alignment
        
        # 데이터 작성

        prev_close_price = None  # 전일 종가 추적용
        
        for row_idx, record in enumerate(backtest_result['daily_records'], 2):
            # 날짜 (첫 데이터와 매주 월요일은 볼드체)
            cell = ws_detail.cell(row=row_idx, column=1, value=record['date'])
            cell.alignment = center_alignment
            
            # 첫 데이터 또는 월요일 체크
            if row_idx == 2:  # 첫 데이터
                cell.font = Font(bold=True)
            else:
                # 날짜에서 요일 추출 (예: "25.01.02.(목)" -> "월")
                date_str = record['date']
                if '(월)' in date_str:
                    cell.font = Font(bold=True)
            
            # 주차
            cell = ws_detail.cell(row=row_idx, column=2, value=record['week'])
            cell.alignment = center_alignment
            
            # RSI
            rsi_value = record.get('rsi', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=3, value=f"{rsi_value:.2f}")
            cell.alignment = center_alignment
            
            # 모드 (SF: 초록색 글자, AG: 주황색 글자)
            cell = ws_detail.cell(row=row_idx, column=4, value=record['mode'])
            cell.alignment = center_alignment
            
            if record['mode'] == 'SF':
                cell.font = Font(color="008000")  # 초록색 글자
            elif record['mode'] == 'AG':
                cell.font = Font(color="FF8C00")  # 주황색 글자
            
            # 현재회차
            cell = ws_detail.cell(row=row_idx, column=5, value=record['current_round'])
            cell.alignment = center_alignment
            
            # 1회시드
            seed_amount = record.get('seed_amount', 0.0) or 0.0
            if seed_amount > 0:
                cell = ws_detail.cell(row=row_idx, column=6, value=f"${seed_amount:,.0f}")
            else:
                cell = ws_detail.cell(row=row_idx, column=6, value="")
            cell.alignment = center_alignment
            
            # 매수주문가
            buy_order_price = record.get('buy_order_price', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=7, value=f"${buy_order_price:.2f}")
            cell.alignment = center_alignment
            
            # 종가 (어제 대비 상승: 빨간색, 하락: 파란색)
            close_price = record.get('close_price', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=8, value=f"{close_price:.2f}")
            cell.alignment = center_alignment
            
            # 전일 대비 상승/하락 색상 적용
            if prev_close_price is not None:
                if close_price > prev_close_price:
                    cell.font = Font(color="FF0000")  # 빨간색
                elif close_price < prev_close_price:
                    cell.font = Font(color="0000FF")  # 파란색
            
            prev_close_price = close_price  # 다음 행을 위해 저장
            
            # 매도목표가
            sell_target_price = record.get('sell_target_price', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=9, value=f"${sell_target_price:.2f}")
            cell.alignment = center_alignment
            
            # 손절예정일
            cell = ws_detail.cell(row=row_idx, column=10, value=record['stop_loss_date'])
            cell.alignment = center_alignment
            
            # 거래일수
            cell = ws_detail.cell(row=row_idx, column=11, value=record['trading_days'])
            cell.alignment = center_alignment
            
            # 매수체결 (빨간색)
            buy_executed_price = record.get('buy_executed_price', 0.0) or 0.0
            if buy_executed_price > 0:
                cell = ws_detail.cell(row=row_idx, column=12, value=f"${buy_executed_price:.2f}")
                cell.font = Font(color="FF0000")  # 빨간색
            else:
                cell = ws_detail.cell(row=row_idx, column=12, value="")
            cell.alignment = center_alignment
            
            # 수량 (매수체결 시 빨간색)
            buy_quantity = record.get('buy_quantity', 0) or 0
            if buy_quantity > 0:
                cell = ws_detail.cell(row=row_idx, column=13, value=buy_quantity)
                cell.font = Font(color="FF0000")  # 빨간색
            else:
                cell = ws_detail.cell(row=row_idx, column=13, value="")
            cell.alignment = center_alignment
            
            # 매수대금 (매수체결 시 빨간색)
            buy_amount = record.get('buy_amount', 0.0) or 0.0
            if buy_amount > 0:
                cell = ws_detail.cell(row=row_idx, column=14, value=f"${buy_amount:,.0f}")
                cell.font = Font(color="FF0000")  # 빨간색
            else:
                cell = ws_detail.cell(row=row_idx, column=14, value="")
            cell.alignment = center_alignment
            
            # 매도일 (파란색 글씨)
            cell = ws_detail.cell(row=row_idx, column=15, value=record['sell_date'])
            cell.alignment = center_alignment
            if record['sell_date']:  # 매도일이 있는 경우에만 파란색 적용
                cell.font = Font(color="0000FF")  # 파란색 글씨
            
            # 매도체결 (파란색 글씨)
            sell_executed_price = record.get('sell_executed_price', 0.0) or 0.0
            if sell_executed_price > 0:
                cell = ws_detail.cell(row=row_idx, column=16, value=f"${sell_executed_price:.2f}")
                cell.font = Font(color="0000FF")  # 파란색 글씨
            else:
                cell = ws_detail.cell(row=row_idx, column=16, value="")
            cell.alignment = center_alignment
            
            # 보유기간
            holding_days = record.get('holding_days', 0) or 0
            if holding_days > 0:
                cell = ws_detail.cell(row=row_idx, column=17, value=f"{holding_days}일")
            else:
                cell = ws_detail.cell(row=row_idx, column=17, value="")
            cell.alignment = center_alignment
            
            # 보유
            cell = ws_detail.cell(row=row_idx, column=18, value=record['holdings'])
            cell.alignment = center_alignment
            
            # 실현손익
            realized_pnl = record.get('realized_pnl', 0.0) or 0.0
            if realized_pnl != 0:
                cell = ws_detail.cell(row=row_idx, column=19, value=f"${realized_pnl:,.0f}")
            else:
                cell = ws_detail.cell(row=row_idx, column=19, value="")
            cell.alignment = center_alignment
            
            # 누적실현
            cumulative_realized = record.get('cumulative_realized', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=20, value=f"${cumulative_realized:,.0f}")
            cell.alignment = center_alignment
            cell.font = Font(color="FF0000")  # 빨간색
            
            # 당일실현
            daily_realized = record.get('daily_realized', 0.0) or 0.0
            if daily_realized != 0:
                cell = ws_detail.cell(row=row_idx, column=21, value=f"${daily_realized:,.0f}")
            else:
                cell = ws_detail.cell(row=row_idx, column=21, value="")
            cell.alignment = center_alignment
            
            # 예수금
            cash_balance = record.get('cash_balance', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=22, value=f"${cash_balance:,.0f}")
            cell.alignment = center_alignment
            
            # 총자산
            total_assets = record.get('total_assets', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=23, value=f"${total_assets:,.0f}")
            cell.alignment = center_alignment
        
        # 열 너비 자동 조정
        for ws in [ws_summary, ws_detail]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        try:
            wb.save(filename)
            print(f"✅ 백테스팅 결과가 엑셀 파일로 저장되었습니다: {filename}")
            return filename
        except Exception as e:
            print(f"❌ 엑셀 파일 저장 실패: {e}")
            return None

def main():
    """메인 실행 함수"""
    print("🚀 SOXL 퀀트투자 시스템")
    print("=" * 50)
    

    # 투자원금 사용자 입력
    while True:
        try:
            initial_capital_input = input("💰 초기 투자금을 입력하세요 (달러): ").strip()
            if not initial_capital_input:
                initial_capital = 9000  # 기본값
                print(f"💰 투자원금: ${initial_capital:,.0f} (기본값)")
                break
            
            initial_capital = float(initial_capital_input)
            if initial_capital <= 0:
                print("❌ 투자금은 0보다 큰 값이어야 합니다.")
                continue
                
            print(f"💰 투자원금: ${initial_capital:,.0f}")
            break
            
        except ValueError:
            print("❌ 올바른 숫자를 입력해주세요.")
            continue
    
    # 트레이더 초기화
    trader = SOXLQuantTrader(initial_capital)
    
    # 시작일 입력(엔터 시 1년 전)
    start_date_input = input("📅 투자 시작일을 입력하세요 (YYYY-MM-DD, 엔터시 1년 전): ").strip()
    if not start_date_input:
        start_date_input = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
    trader.session_start_date = start_date_input
    
    while True:
        print("\n" + "=" * 50)
        print("메뉴를 선택하세요:")
        print("1. 오늘의 매매 추천 보기")
        print("2. 포트폴리오 현황 보기")
        print("3. 백테스팅 실행")
        print("4. 매수 실행 (테스트)")
        print("5. 매도 실행 (테스트)")
        print("T. 테스트 날짜(오늘) 설정/해제")
        print("6. 종료")
        
        choice = input("\n선택 (1-6): ").strip()
        
        if choice == '1':
            # 저장된 시작일부터 오늘까지 시뮬레이션으로 현재 상태 산출
            start_date = trader.session_start_date or (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
            sim_result = trader.simulate_from_start_to_today(start_date, quiet=True)
            if "error" in sim_result:
                print(f"❌ 시뮬레이션 실패: {sim_result['error']}")
            
            # 현재 상태 기반 오늘의 추천 출력
            recommendation = trader.get_daily_recommendation()
            trader.print_recommendation(recommendation)
            
        elif choice == '2':
            # 저장된 시작일부터 오늘까지 시뮬레이션으로 현황 재계산
            start_date = trader.session_start_date or (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
            sim_result = trader.simulate_from_start_to_today(start_date, quiet=True)
            if "error" in sim_result:
                print(f"❌ 시뮬레이션 실패: {sim_result['error']}")
            
            # 기존 형식 유지하여 현황 출력
            if trader.positions:
                print("\n💼 현재 포트폴리오:")
                print("-" * 30)
                for pos in trader.positions:
                    hold_days = (datetime.now() - pos['buy_date']).days
                    print(f"{pos['round']}회차: {pos['shares']}주 @ ${pos['buy_price']:.2f} ({hold_days}일)")
                print(f"\n현금잔고: ${trader.available_cash:,.0f}")
            else:
                print("\n보유 포지션이 없습니다.")
                print(f"현금잔고: ${trader.available_cash:,.0f}")
        
        elif choice == '3':
            # 백테스팅 실행
            print("\n📊 백테스팅 실행")
            print("-" * 30)
            
            start_date = input("시작 날짜를 입력하세요 (YYYY-MM-DD): ").strip()
            if not start_date:
                print("날짜를 입력해주세요.")
                continue
            
            end_date = input("종료 날짜를 입력하세요 (YYYY-MM-DD, 엔터시 오늘까지): ").strip()
            if not end_date:
                end_date = None
            
            print("\n백테스팅을 시작합니다...")
            backtest_result = trader.run_backtest(start_date, end_date)
            
            if "error" in backtest_result:
                print(f"❌ 백테스팅 실패: {backtest_result['error']}")
                continue
            

            # MDD 계산
            mdd_info = trader.calculate_mdd(backtest_result['daily_records'])
            
            # 결과 출력
            print("\n" + "=" * 60)
            print("📊 백테스팅 결과 요약")
            print("=" * 60)
            print(f"기간: {backtest_result['start_date']} ~ {backtest_result['end_date']}")
            print(f"거래일수: {backtest_result['trading_days']}일")
            print(f"초기자본: ${backtest_result['initial_capital']:,.0f}")
            print(f"최종자산: ${backtest_result['final_value']:,.0f}")
            print(f"총수익률: {backtest_result['total_return']:+.2f}%")

            print(f"최대 MDD: {mdd_info.get('mdd_percent', 0.0):.2f}%")
            print(f"최종보유포지션: {backtest_result['final_positions']}개")

            print(f"총 거래일수: {len(backtest_result['daily_records'])}일")
            
            # 엑셀 내보내기 여부 확인
            export_choice = input("\n엑셀 파일로 내보내시겠습니까? (y/n): ").strip().lower()
            if export_choice == 'y':
                filename = trader.export_backtest_to_excel(backtest_result)
                if filename:
                    print(f"📁 파일 위치: {os.path.abspath(filename)}")
            
        elif choice == '4':
            print("\n🔧 매수 테스트 기능 (개발 중)")
            
        elif choice == '5':
            print("\n🔧 매도 테스트 기능 (개발 중)")
            
        elif choice.lower() == 't':
            print("\n🧪 테스트 날짜 설정")
            print("- 비우고 엔터하면 해제됩니다")
            test_date = input("테스트 오늘 날짜 (YYYY-MM-DD): ").strip()
            trader.set_test_today(test_date if test_date else None)
            
        elif choice == '6':
            print("프로그램을 종료합니다.")
            break
            
        else:
            print("올바른 선택지를 입력하세요.")

if __name__ == "__main__":
    main()

